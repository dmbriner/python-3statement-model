from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .model import ModelOutput


def _format_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="0F4C81")
    header_font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(len(max(values, key=len, default="")) + 2, 12), 36)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _write_sheets(
    writer: pd.ExcelWriter,
    output: ModelOutput,
    sensitivity: pd.DataFrame,
    historical: pd.DataFrame,
    valuation_summary: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
) -> None:
    historical.to_excel(writer, sheet_name="Historical", index=False)
    output.income_statement.to_excel(writer, sheet_name="Income Statement", index=False)
    output.balance_sheet.to_excel(writer, sheet_name="Balance Sheet", index=False)
    output.cash_flow.to_excel(writer, sheet_name="Cash Flow", index=False)
    output.fcf.to_excel(writer, sheet_name="FCF", index=False)
    output.ppe_schedule.to_excel(writer, sheet_name="PP&E Schedule", index=False)
    output.debt_schedule.to_excel(writer, sheet_name="Debt Schedule", index=False)
    output.equity_schedule.to_excel(writer, sheet_name="Equity Schedule", index=False)
    sensitivity.to_excel(writer, sheet_name="Sensitivity")
    if valuation_summary is not None and not valuation_summary.empty:
        valuation_summary.to_excel(writer, sheet_name="Valuation", index=False)
    for sheet_name, df in (additional_sheets or {}).items():
        safe_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=safe_name, index=False)

    for worksheet in writer.book.worksheets:
        _format_sheet(worksheet)


def build_excel_bytes(
    output: ModelOutput,
    sensitivity: pd.DataFrame,
    historical: pd.DataFrame,
    valuation_summary: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    """Return a multi-sheet Excel workbook as bytes (for st.download_button)."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheets(writer, output, sensitivity, historical, valuation_summary, additional_sheets)
    return buffer.getvalue()


def export_model_to_excel(
    output: ModelOutput,
    sensitivity: pd.DataFrame,
    historical: pd.DataFrame,
    out_path: str | Path,
    valuation_summary: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _write_sheets(writer, output, sensitivity, historical, valuation_summary, additional_sheets)

    return out_path
